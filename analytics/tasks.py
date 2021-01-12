from analytics.models import *
from openpyxl import load_workbook
from django.core.exceptions import ObjectDoesNotExist


def download_to_base(request):
    """Получаем Excel файл из формы и заполняем базу тегами"""

    if request.method == 'POST':
        uploaded_file = request.FILES['excel-file']  # получаем содержимое из формы загрузки файлов "name="excel-file""
        wb = load_workbook(filename=uploaded_file, read_only=False)  # получаем excel файл
        ws = wb.active  # делаем лист активным в памяти

    """вычитываем имя группы тегов из каждой строки, пытаемся извлечь из базы имя группы
    соответствующее текущей строчке тега, если имя группы есть - пишем в базу соответствующие теги, привязанные к этой
    группе соотношением один-комногим, если этого имени нет выпадает исключение. Обрабатываем сохраняя
    имя нруппы в таблицу Group и пишем в базу соответствующие теги, привязанные к этой 
    группе соотношением один-комногим
    Если файл с тегами не содержит нужные данные (пустые колонки или ячейки (кроме комментариев)) вывести уведомление 
    об этом"""
    for i in range(1, ws.max_row+1):
        name_group = (ws.cell(row=i, column=6)).value  # присваиваем переменной name_group значение ячейки столбца group
        try:
            Group.objects.get(name_group=name_group)  # получаем из базы значение равное переменной name_group
        except ObjectDoesNotExist:  # исключение если такого значения в базе нет
            save_set_of_groups = Group(name_group=name_group)  # сохраняем содержимое переменной name_group в базу
            save_set_of_groups.save()
        group = Group.objects.get(name_group=name_group)  # заносим в переменную объект группы для текущей записи тега
                                                            # чтобы создать связь один ко многим с таблицей Tags
        name_tag = (ws.cell(row=i, column=1)).value  # выбираем ячейки из таблицы
        if not isinstance(name_tag, (str, int, bool, float)):
            continue
        address_tag = (ws.cell(row=i, column=4)).value
        name_tag = str(name_tag)
        address_tag = str(address_tag)
        try:
            Tags.objects.get(address=address_tag)
        except ObjectDoesNotExist:
            comment_tag = (ws.cell(row=i, column=5)).value
            comment_tag = str(comment_tag)
            """"Для случая когда теги импортированы из таблицы тегов"""
            if address_tag.startswith('%'):
                if comment_tag.capitalize == 'Spare':
                    continue
                tag_table = (ws.cell(row=i, column=2)).value
                data_type = (ws.cell(row=i, column=3)).value
                tag_table = str(tag_table)
                data_type = str(data_type)
            else:
                """"Для случая когда теги полуены из DB блоков"""
                data_type = (ws.cell(row=i, column=2)).value
                tag_table = (ws.cell(row=i, column=4)).value
                if name_tag == 'InOut' or name_tag == 'Input' or name_tag == 'Output' or name_tag == 'Spare' or name_tag == 'Static' or data_type == 'Struct':
                    continue
                if data_type == 'Bool':
                    x = '.DBX'
                elif data_type == 'Real' or data_type == 'Time_Of_Day' or data_type == 'Date_And_Time' or data_type == 'DInt' or data_type == 'IEC_TIMER':
                    x = '.DBD'
                elif data_type == 'Int' or 'Word':
                    x = '.DBW'
                address_tag = f'{(ws.cell(row=i, column=4)).value}{x}{(ws.cell(row=i, column=3)).value}'

            save_set_of_tags = Tags(group=group, name_tag=name_tag, tag_table=tag_table, data_type=data_type,
                                    address=address_tag, comment=comment_tag)
            save_set_of_tags.save()      # сохраняем модель Tags

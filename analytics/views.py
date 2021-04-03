# # Create your views here.
from django.http import JsonResponse
import plotly.graph_objects as go
from analytics.models import *
from django.shortcuts import render
from openpyxl import load_workbook
from django.core.exceptions import ObjectDoesNotExist
from influxdb import InfluxDBClient
# from celery import task
from analytics.tasks import *


# @task
def pars_tags_list(request):
    """Получаем Excel файл из формы и заполняем базу тегами"""
    download_to_base(request)
    # if request.method == 'POST':
    #     uploaded_file = request.FILES['excel-file']  # получаем содержимое из формы загрузки файлов "name="excel-file""
    #     wb = load_workbook(filename=uploaded_file, read_only=False)  # получаем excel файл
    #     ws = wb.active  # делаем лист активным в памяти

    """вычитываем имя группы тегов из каждой строки, пытаемся извлечь из базы имя группы
    соответствующее текущей строчке тега, если имя группы есть - пишем в базу соответствующие теги, привязанные к этой
    группе соотношением один-комногим, если этого имени нет выпадает исключение. Обрабатываем сохраняя
    имя нруппы в таблицу Group и пишем в базу соответствующие теги, привязанные к этой 
    группе соотношением один-комногим
    Если файл с тегами не содержит нужные данные (пустые колонки или ячейки (кроме комментариев)) вывести уведомление 
    об этом"""
    # for i in range(1, ws.max_row+1):
    #     name_group = (ws.cell(row=i, column=6)).value  # присваиваем переменной name_group значение ячейки столбца group
    #     try:
    #         Group.objects.get(name_group=name_group)  # получаем из базы значение равное переменной name_group
    #     except ObjectDoesNotExist:  # исключение если такого значения в базе нет
    #         save_set_of_groups = Group(name_group=name_group)  # сохраняем содержимое переменной name_group в базу
    #         save_set_of_groups.save()
    #     group = Group.objects.get(name_group=name_group)  # заносим в переменную объект группы для текущей записи тега
    #                                                         # чтобы создать связь один ко многим с таблицей Tags
    #     name_tag = (ws.cell(row=i, column=1)).value  # выбираем ячейки из таблицы
    #     if not isinstance(name_tag, (str, int, bool, float)):
    #         continue
    #     address_tag = (ws.cell(row=i, column=4)).value
    #     name_tag = str(name_tag)
    #     address_tag = str(address_tag)
    #     try:
    #         Tags.objects.get(address=address_tag)
    #     except ObjectDoesNotExist:
    #         comment_tag = (ws.cell(row=i, column=5)).value
    #         comment_tag = str(comment_tag)
    #         """"Для случая когда теги импортированы из таблицы тегов"""
    #         if address_tag.startswith('%'):
    #             if comment_tag.capitalize == 'Spare':
    #                 continue
    #             tag_table = (ws.cell(row=i, column=2)).value
    #             data_type = (ws.cell(row=i, column=3)).value
    #             tag_table = str(tag_table)
    #             data_type = str(data_type)
    #         else:
    #             """"Для случая когда теги полуены из DB блоков"""
    #             data_type = (ws.cell(row=i, column=2)).value
    #             tag_table = (ws.cell(row=i, column=4)).value
    #             if name_tag == 'InOut' or name_tag == 'Input' or name_tag == 'Output' or name_tag == 'Spare' or name_tag == 'Static' or data_type == 'Struct':
    #                 continue
    #             if data_type == 'Bool':
    #                 x = '.DBX'
    #             elif data_type == 'Real' or data_type == 'Time_Of_Day' or data_type == 'Date_And_Time' or data_type == 'DInt' or data_type == 'IEC_TIMER':
    #                 x = '.DBD'
    #             elif data_type == 'Int' or 'Word':
    #                 x = '.DBW'
    #             address_tag = f'{(ws.cell(row=i, column=4)).value}{x}{(ws.cell(row=i, column=3)).value}'
    #
    #         save_set_of_tags = Tags(group=group, name_tag=name_tag, tag_table=tag_table, data_type=data_type,
    #                                 address=address_tag, comment=comment_tag)
    #         save_set_of_tags.save()      # сохраняем модель Tags

    return render(request, 'settings/settings.html')


def group_prepare(request):
    """Принимаем AJAX от клиента с названием группы по которой нужно вывести список тегов и комментариев
    принадлежащих к этой группе, фильтруем и создаем словарь с парами имя - комментарий, присвиваем в data и отправляем
    обратно JsonResponse(data)"""
    if request.method == 'POST':
        data = request.POST
        tags_comments_list = {}
        for e in Tags.objects.filter(group=Group.objects.get(name_group=data.get('groupList'))):
            tags_comments_list[e.name_tag] = e.comment, e.data_type
            print(e)
        data = tags_comments_list
        return JsonResponse(data)


def tags_influx_prepare(request):
    """Подготовка списка переменных и тегов для тренда"""

    if request.method == 'POST':
        influx_data = request.POST
        print('influx_data ', influx_data)
        influx_query_tags = influx_data.get('list')
        print(influx_query_tags)
        list_influx_query_tags = influx_query_tags.split(',')
        list_influx_query_tags.append('time')
        influx_query_tags = ','.join('"{0}"'.format(w) for w in influx_query_tags.split(','))  # генератор списка тегов
        print(influx_query_tags)
        time_before = influx_data.get('timeClientUtcValueFrom')  # timeFrom в UTC
        time_after = influx_data.get('timeClientUtcValueTo')  # timeTo в UTC

        bucket = "CNC"  # Имя базы данных
        measurement = "line"  # Имя измерения

        client = InfluxDBClient('192.168.2.163', 8086, 'root', 'root')  # Подключение к базе. В будущем нужно сделать шаблон
        client.create_database(bucket)
        # внесения настроек подключения к базе. Создать модель с настройками и делать из нее выборку
        list_database = client.get_list_database()  # Список баз данных
        client.switch_database(bucket)  # Переключение на нужную базу

        query = f'SELECT {influx_query_tags} FROM {bucket}."autogen".{measurement} WHERE time >= \'{time_before}\' AND time < \'{time_after}\''  # Запрос в Influx
        print(query)
        result = client.query(query).get_points()
        s = []
        result_response = {}
        for items in result:
            s.append(items)

        for element in list_influx_query_tags:
            item = []
            for i in s:
                if str(i[element]) == 'True':
                    i[element] = 1
                elif str(i[element]) == 'False':
                    i[element] = 0
                item.append(i[element])
            result_response[element] = item
        s.clear()
        print(result_response)
        response = {'result': result_response}
        return JsonResponse(response, safe=False)


def analytics_render(request):
    """Отрисовка страницы"""
    group = Group.objects.all()
    tags = Tags.objects.all()
    comment = Tags.objects.all()
    return render(request, 'analytics/analytics.html', context={'plot_div': chart(), 'group': group, 'tags': tags,
                                                                'comment': comment}, )


def chart():
    """Создание тренда"""
    my_query_1 = [1, 10, 1, 10, 1, 10, 1, 10, 1, 10, 1, 10, 1, 10, 1, 10, 1, 10, 1]
    my_query_2 = [1, 2, 3, 4, 5, 6, 7, 8, 9, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    # генератор точек для оси Х в соответствии с количеством точек оси У
    q = [i for i in range(1, len(my_query_1))]

    # рисуем график
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=q, y=my_query_1,  # заполнение осей значениями
        name="yaxis1 data",  # лейбл
        yaxis="y1",
        fillcolor="#1f77b4"
    ))

    fig.add_trace(go.Scatter(
        x=q, y=my_query_2,
        name="yaxis2 data",
        yaxis="y2",
        fillcolor="#ff7f0e"
    ))
    #     #
    #     # fig.add_trace(go.Scatter(
    #     #     x=q, y=[0, 1, 0, 1, 0, 1, 0],
    #     #     name="yaxis3 data",
    #     #     yaxis="y3",
    #     #     fillcolor="#d62728"
    #     # ))
    #     #
    #     # fig.add_trace(go.Scatter(
    #     #     x=q, y=[0, 1, 0, 1, 0, 1, 0],
    #     #     name="yaxis4 data",
    #     #     yaxis="y4",
    #     #     fillcolor="#9467bd"
    #     # ))
    #
    # Create axis objects
    fig.update_layout(
        xaxis=dict(domain=[0.15, 1], rangeslider=dict(visible=True)),

        yaxis1=dict(title="yaxis1 title", titlefont=dict(color="#1f77b4"), tickfont=dict(color="#1f77b4"),
                    anchor="free", side="left", position=0.15),

        yaxis2=dict(title="yaxis2 title", titlefont=dict(color="#ff7f0e"), tickfont=dict(color="#ff7f0e"),
                    anchor="free", overlaying="y", side="left", position=0.1),
        #         #
        #         # yaxis3=dict(title="yaxis3 title", titlefont=dict(color="#d62728"), tickfont=dict(color="#d62728"),
        #         #             anchor="free", overlaying="y", side="left", position=0.05),
        #         #
        #         # yaxis4=dict(title="yaxis4 title", titlefont=dict(color="#9467bd"), tickfont=dict(color="#9467bd"),
        #         #             anchor="free", overlaying="y", side="left", position=0)
    )

    # Update layout properties
    fig.update_layout(title_text="multiple y-axes example", width=1800, )

    return fig.to_html('plot_div')
